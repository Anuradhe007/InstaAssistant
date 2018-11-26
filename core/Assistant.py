import instaloader
import datetime
from itertools import dropwhile, takewhile
import schedule
import time
from tkinter import messagebox
import sys
from openpyxl import Workbook
import math
import random
import requests

class Assistant:

    def initializeAssistant(self, userName, password, userNameList, timeToSearch, proxy, filePath):
        self.userName = userName
        self.password = password
        self.userNameList = userNameList
        self.timeToSearch = int(timeToSearch)
        self.proxy = proxy
        self.filePath = filePath
        self.startedTime = datetime.datetime.strptime(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                                      "%Y-%m-%d %H:%M:%S")
        self.endingTime = datetime.datetime.strptime(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                                     "%Y-%m-%d %H:%M:%S") + datetime.timedelta(hours=24)
        self.s = requests.Session()
        self.schedulerInitialize()

    def getUserPosts(self, userName):
        loadrObject = instaloader.Instaloader()
        try:
            self.setProxy(self.proxy)
            loadrObject.login(self.userName, self.password)
            time.sleep(random.randint(2, 10))
        except:
            print('Login error for username -> ' + self.userName)
            sys.exit()
        self.loaderObject = loadrObject

        posts = instaloader.Profile.from_username(self.loaderObject.context, userName).get_posts()
        self.loaderObject.close()
        time.sleep(random.randint(30, 120))

        SINCE = self.startedTime
        UNTIL = self.endingTime
        filteredPosts = []
        for post in posts:
            if SINCE <= post.date_local <= UNTIL:
                filteredPosts.append(post)
        return filteredPosts

    def getPostLikers(self, post):
        likers = []
        for like in post.get_likes():
            likers.append(like.username)
        return likers

    def getUserFollowers(self, userName):
        followers = instaloader.Profile.from_username(self.loaderObject.context, userName).get_followers()
        time.sleep(random.randint(30, 120))
        return followers

    def extractNecessaryDetails(self, userNameList):
        for userName in userNameList:
            print(userName)
            userPosts = self.getUserPosts(userName)
            for post in userPosts:
                likers = self.getPostLikers(post)
                details = [('User name', 'Post created date time', 'Caption', 'Total likes', 'Likers'),
                           (userName, post.date_local, post.caption, len(likers), '')]
                for liker in likers:
                    details.append(('', '', '', '', liker))
                    self.createExcelFile(userName, post.date_local, details, self.filePath)

    def job(self):
        self.extractNecessaryDetails(self.userNameList)
        print('Executed the job!')
        time.sleep(random.randint(2, 5))

    def jobScheduler(self, timeToRun, job):
        print('Scheduler started!')
        # schedule.every(timeToRun).seconds.do(job)
        schedule.every(timeToRun).minutes.do(job)
        # schedule.every().hour.do(job)
        # schedule.every().day.at("10:30").do(job)

        count = 0
        while 1:
            schedule.run_pending()
            time.sleep(1)
            if self.endingTime <= datetime.datetime.now():
                self.generateFollowers()
                messagebox.showinfo("Info", "Finished Data Collecting")
                break
            count = count + 1

    def schedulerInitialize(self):
        self.jobScheduler(self.timeToSearch, self.job)

    def job_that_executes_once(self):
        return schedule.CancelJob

    def generateFollowers(self):
        print("Followers generator")

        for userName in self.userNameList:
            loadrObject = instaloader.Instaloader()
            try:
                loadrObject.login(self.userName, self.password)
                time.sleep(random.randint(5, 10))
            except:
                print('Login error for username -> ' + self.userName)
                sys.exit()
            self.loaderObject = loadrObject
            followers = self.getUserFollowers(userName)
            self.writeFollowersToExcel(userName, followers, self.filePath)
            self.loaderObject.close()
            time.sleep(random.randint(30, 120))

    def createExcelFile(self, userName, postCreatedTime, detailsList, filePath):
        fileName = self.fileNameCreator(userName, postCreatedTime)
        workbook = Workbook()
        worksheet = workbook.active

        for data in detailsList:
            worksheet.append(data)
        workbook.save(filePath + fileName)

    def fileNameCreator(self, userName, postCreationTime):
        todayDate = datetime.datetime.now()
        spentTimeInMinutes = math.ceil((datetime.datetime.strptime(todayDate.strftime("%Y-%m-%d %H:%M:%S"),
                                                                   "%Y-%m-%d %H:%M:%S") - postCreationTime).seconds / 60)
        hours = math.floor(spentTimeInMinutes / 60)

        fileName = userName + '-' + str(postCreationTime.strftime("%Y-%m-%d %H-%M-%S")) + '-Likes After '
        if hours > 0:
            fileName = fileName + str(hours) + ' hours ' + str(
                spentTimeInMinutes % 60) + ' minutes_' + todayDate.strftime('%b %d,%Y') + '.xlsx'
        else:
            fileName = fileName + str(spentTimeInMinutes) + 'minutes_' + todayDate.strftime('%b %d,%Y') + '.xlsx'
        return fileName

    def writeFollowersToExcel(self, userName, followers, filePath):
        todayDate = datetime.datetime.now()
        fileName = userName + '-' + 'Followers for last 24 hrs-' + todayDate.strftime('%b %d,%Y') + '.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        count = 1
        for follower in followers:
            worksheet.cell(row=count, column=1, value=follower.username)
            count = count + 1
        workbook.save(filePath + fileName)

    def setProxy(self, proxy=None):
        """
        Set proxy for all requests::
        Proxy format - user:password@ip:port
        """
        if proxy is not None:
            proxies = {'http': proxy, 'https': proxy}
            self.s.proxies.update(proxies)
